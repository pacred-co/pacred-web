# -*- coding: utf-8 -*-
"""Build the Pacred ใบกำกับภาษี (tax-invoice) form — 4 roles: CS → Pricing → Docs → Account.
Faithful to the AXELRA template styling; adds the missing PRICING (cost) section;
encodes the 3-number model: SELLING (CS→VAT) / COST (Pricing→PEAK+profit) / สำแดง (Docs→ใบขน)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment

FONT = "Bai Jamjuree"
# palette (from the original)
RED="E06666"; PINK="F4CCCC"; GREY="EFEFEF"; BLUE="073763"; CYAN="00FFFF"
LBLUE="C9DAF8"; MBLUE="6D9EEB"; GREEN="274E13"; LGREEN="D9EAD3"; TEAL="D0E0E3"
GOLD="BF9000"; SBLUE="4472C4"; SRED="FF0000"; MAROON="990000"; LPINK2="EAD1DC"
PURPLE="674EA7"; LPUR="D9D2E9"; YEL="FFFF00"; ORANGE="E69138"; LORANGE="FCE5CD"; WHITE="FFFFFF"

thin = Side(style="thin", color="999999")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def F(color="000000", b=False, sz=12, name=FONT):
    return Font(name=name, size=sz, bold=b, color=color)
def Fill(c): return PatternFill("solid", fgColor=c)
AC = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL = Alignment(horizontal="left",   vertical="center", wrap_text=True)
AR = Alignment(horizontal="right",  vertical="center", wrap_text=True)

def put(ws, coord, val=None, fill=None, font=None, align=AC, border=True):
    c = ws[coord]
    if val is not None: c.value = val
    if fill: c.fill = Fill(fill)
    c.font = font or F()
    c.alignment = align
    if border: c.border = BORDER
    return c

def merge(ws, rng): ws.merge_cells(rng)

wb = Workbook()
ws = wb.active
ws.title = "ฟอร์มใบกำกับภาษี"
ws.sheet_view.showGridLines = False

# ── column widths (match original + extend) ──
W = {"A":9.4,"B":14.2,"C":9.9,"D":9,"E":12.5,"F":10.6,"G":11,"H":9,"I":18,"J":9,
     "K":9,"L":15.6,"M":9,"N":10,"O":11,"P":11,"Q":11,"R":11,"S":13,"T":11,"U":11,
     "V":10,"W":13,"X":11}
for k,v in W.items(): ws.column_dimensions[k].width = v
for col in "YZ": ws.column_dimensions[col].width = 11
for col in ["AA","AB","AC","AD","AE","AF","AG","AH","AI","AJ","AK","AL","AM","AN","AO","AP","AQ"]:
    ws.column_dimensions[col].width = 9

# ════════════════════════════════════════════════════════════════════
# SECTION 1 — CS (ผู้ดูแลลูกค้า) : customer + product + SELLING price + VAT
# ════════════════════════════════════════════════════════════════════
ws.row_dimensions[1].height=40.5; ws.row_dimensions[2].height=34
merge(ws,"A1:X1"); put(ws,"A1","AXELRA (THAILAND) CO., LTD.",RED,F(WHITE,True,23))
merge(ws,"A2:X2"); put(ws,"A2","ADD : 14 SOI PHET KASEM 77 YAEK 3-6, NONG KHANG PHLU, NONG KHAEM, BANGKOK 10160 · TAX ID 0105564077716 · TEL 02-421-3325",RED,F(WHITE,False,12))
merge(ws,"Y1:AQ1"); put(ws,"Y1","แนบใบสลิปการโอนเงิน ไทย–จีน",GOLD,F(YEL,True,18))

# left info block (labels pink, values grey)
def info(ws, lr, label, vr, val=None, lab_fill=PINK, val_fill=GREY, span=None):
    put(ws, lr, label, lab_fill, F("434343",True,12), AL)
    if span: merge(ws, span)
    put(ws, vr, val, val_fill, F("000000",False,12), AL)

ws.row_dimensions[3].height=26
info(ws,"A3","ชื่อ บริษัทฯ / บุคคล","C3",span="C3:F3")
info(ws,"A4","ที่อยู่","C4",span="C4:F5"); merge(ws,"A4:A5")
info(ws,"A6","เลขทะเบียน 13 หลัก","C6",span="C6:F6")
info(ws,"A7","สำนักงานใหญ่ / สาขา","C7",span="C7:F7")
info(ws,"A8","โอนผ่านบัญชี","C8","225-2-91144-0 (ธ.กสิกรไทย)",span="C8:F8")
info(ws,"A9","รหัสลูกค้า","C9",span="C9:F9")
info(ws,"A10","เลขที่ปิดตู้","C10",span="C10:F10")
# mid contact block
info(ws,"G3","ข้อมูลผู้ติดต่อ","J3",span="J3:M3"); merge(ws,"G3:I3")
info(ws,"G4","เบอร์โทรศัพท์","J4",span="J4:M4"); merge(ws,"G4:I4")
info(ws,"G5","Email","J5",span="J5:M5"); merge(ws,"G5:I5")
put(ws,"G6","เรทหยวน\n(ขาย)",BLUE,F(CYAN,True,11))
put(ws,"H6",4.0,CYAN,F("000000",True,14)); ws["H6"].number_format="0.00"
merge(ws,"I6:M6"); put(ws,"I6","← เรทหยวนที่ใช้คิดราคาขายลูกค้า (THB/¥)",None,F("999999",False,9),AL)
# right doc block
info(ws,"N3","วันที่ ออกใบกำกับภาษี","R3",span="R3:X4",lab_fill=PINK); merge(ws,"N3:Q4")
info(ws,"N5","วันที่ ชำระเงิน","R5",span="R5:X5"); merge(ws,"N5:Q5")
put(ws,"R5",'*วันที่ลูกค้าโอนเงิน หากมี 2 ยอด ใส่ 2 ตัวอย่าง (01/12/2026 - 02/12/2026)',GREY,F("CC0000",False,9),AL)
info(ws,"N6","เลขที่เอกสารใบกำกับ","R6",span="R6:X6"); merge(ws,"N6:Q6")
info(ws,"N7","Link Download","R7",span="R7:X7"); merge(ws,"N7:Q7")
info(ws,"N8","IMPORT CODE","R8",span="R8:X8"); merge(ws,"N8:Q8")
info(ws,"N9","SHIPMENT NO","R9",span="R9:X9"); merge(ws,"N9:Q9")
info(ws,"N10","วันที่ / ตู้ถึงไทย","R10",span="R10:X10"); merge(ws,"N10:Q10")
# doc-mode toggle (the cargo/VAT switch) — placed in the free mid-block (rows 7-10)
merge(ws,"G7:I7"); put(ws,"G7","รูปแบบเอกสาร",GOLD,F(WHITE,True,11))
merge(ws,"J7:M7"); put(ws,"J7","เอาเอกสาร (VAT7%)",LORANGE,F("434343",True,11))
merge(ws,"G8:M10"); put(ws,"G8","เอาเอกสาร = ลูกค้าชำระ +VAT7% → ได้ใบกำกับภาษี (งาน CARGO ออกใบขนรวมชื่อ Pacred · ลูกค้าไม่เห็นใบขน)\nไม่เอาเอกสาร = ไม่มี VAT · ไม่มีเอกสาร · ได้ของอย่างเดียว\nอยากได้ใบขนชื่อตัวเอง = กลายเป็นงาน FREIGHT LCL",LORANGE,F("434343",False,9),AL)

# slip image areas (right)
merge(ws,"AA6:AF7"); put(ws,"AA6","ราคาขาย / รายรับ",SBLUE,F(WHITE,False,18))
merge(ws,"AJ6:AO7"); put(ws,"AJ6","ราคาซื้อ / รายจ่าย",SRED,F(WHITE,False,18))
merge(ws,"AA9:AF9"); put(ws,"AA9","สลิปโอนเงินไทย (ยอดรวม VAT)",SBLUE,F(WHITE,True,11))
merge(ws,"AJ9:AO9"); put(ws,"AJ9","สลิปโอนเงินจีน (ต้นทุน · Pricing แนบ)",SRED,F(WHITE,True,11))
merge(ws,"Z10:AG24"); put(ws,"Z10","[ วางรูปสลิปโอนเงินไทย ]",WHITE,F("999999",False,11))
merge(ws,"AI10:AP24"); put(ws,"AI10","[ วางรูปสลิปโอนเงินจีน ]",WHITE,F("999999",False,11))

# product table header (row 11) — col fills: BLUE for index/total, RED for product/price
ws.row_dimensions[11].height=58
hdr1 = {"A11":("ลำดับ",BLUE),"B11":("รูปสินค้า",BLUE),"C11":("HS CODE\n(พิกัดสินค้า)",RED),
 "E11":("เรทอากร\nFE %",RED),"F11":("รหัสสินค้า\n(ซื้อซ้ำ)",RED),"G11":("ชื่อสินค้า (อังกฤษ)",RED),
 "J11":("ชื่อสินค้า (ไทย)",RED),"L11":("BRAND",RED),"M11":("จำนวน",RED),"N11":("หน่วย",RED),
 "O11":("RMB/PCS\nราคาขายหยวน\nต่อชิ้น",RED),"P11":("THB/PCS\nราคาขายบาท\nต่อชิ้น",BLUE),
 "Q11":("ราคา–อากร\n(ฐานสำแดง)",BLUE),"R11":("TOTAL THB\n(ราคาขาย)",BLUE),"S11":("แทร็คกิ้ง\nจัดส่งจากจีน",BLUE),
 "T11":("จำนวนกล่อง\nเข้าโกดัง",BLUE),"U11":("น้ำหนักรวม\nKG.",BLUE),"V11":("ถอด\nอากร",BLUE),
 "W11":("หมายเหตุ\nงาน/สินค้า",BLUE),"X11":("ขนส่ง",BLUE)}
merge(ws,"C11:D11"); merge(ws,"G11:I11"); merge(ws,"J11:K11")
for c,(t,fl) in hdr1.items(): put(ws,c,t,fl,F(WHITE,True,11))

# 20 product rows (12-31)
for i in range(20):
    r=12+i
    ws.row_dimensions[r].height=34
    merge(ws,f"C{r}:D{r}"); merge(ws,f"G{r}:I{r}"); merge(ws,f"J{r}:K{r}")
    put(ws,f"A{r}",i+1,LBLUE,F("434343"))
    for col in ["B","C","E","F","G","J","L","M","O","S","T","U","W","X"]:
        put(ws,f"{col}{r}",None,GREY,F("434343"),AC)
    put(ws,f"N{r}","PCS.",GREY,F("434343"))
    put(ws,f"P{r}",f"=O{r}*$H$6",None,F("000000")); ws[f"P{r}"].number_format="#,##0.0000"
    put(ws,f"Q{r}",f'=IFERROR(P{r}/(1+IF(E{r}>=1,E{r}/100,E{r})),"")',None,F("000000")); ws[f"Q{r}"].number_format="#,##0.0000"
    put(ws,f"R{r}",f"=M{r}*P{r}",None,F("000000",True)); ws[f"R{r}"].number_format="#,##0.00"
    # duty amount on the ex-duty base (fixed: use normalized rate, not raw E)
    put(ws,f"V{r}",f"=ROUND(Q{r}*IF(E{r}>=1,E{r}/100,E{r})*M{r},2)",None,F("000000")); ws[f"V{r}"].number_format="#,##0.00"
    ws[f"O{r}"].number_format="#,##0.0000"; ws[f"E{r}"].number_format='0.##"%"'

# totals (32-35)
for r,lab,fill,fml,fl2 in [(32,"ยอดสุทธิ (ราคาขาย)",MBLUE,"=SUM(R12:R31)",MBLUE),
                           (33,"ภาษีมูลค่าเพิ่ม 7%",MBLUE,"=ROUND(R32*0.07,2)",MBLUE),
                           (34,"อากรที่ต้องจ่ายเพิ่ม",MBLUE,"=SUM(V12:V31)",MBLUE),
                           (35,"รวมทั้งสิ้น (ลูกค้าชำระ)",BLUE,"=R32+R33",BLUE)]:
    merge(ws,f"N{r}:Q{r}"); put(ws,f"N{r}",lab,fill,F(WHITE,True,13),AL)
    put(ws,f"R{r}",fml,fl2,F(WHITE,True,12)); ws[f"R{r}"].number_format="#,##0.00"
merge(ws,"T32:U32"); put(ws,"T32","=SUM(T12:T31)",MBLUE,F(WHITE,True,11)); ws["T32"].number_format="#,##0"
put(ws,"A32","หมายเหตุ :: ใช้ทศนิยมไม่เกิน 4 ตำแหน่ง",None,F("CC0000",False,10),AL,border=False); merge(ws,"A32:I32")

print("section 1 done")

def bar(ws, r, rng, text, fill):
    merge(ws, rng); ws.row_dimensions[r].height=46
    put(ws, rng.split(":")[0], text, fill, F(WHITE,True,20), AL)

# ════════════════════════════════════════════════════════════════════
# SECTION 2 — PRICING (เจ้าหน้าที่ Pricing) : COST price + China slip + profit
# ════════════════════════════════════════════════════════════════════
bar(ws,37,"A37:X37","2.  สำหรับเจ้าหน้าที่ Pricing เท่านั้น  —  ราคาต้นทุน (จีน) + แนบสลิปจีน",MAROON)
put(ws,"A38","เรทหยวน\nต้นทุน",MAROON,F(WHITE,True,11))
put(ws,"B38","=H6",GREY,F("000000",True,13)); ws["B38"].number_format="0.00"
merge(ws,"C38:X38"); put(ws,"C38","เรทหยวนที่ Pricing จ่ายซื้อจีนจริง (THB/¥) · ค่าเริ่มต้น = เรทขาย แก้ได้ถ้าจ่ายคนละเรท · ต้นทุนนี้ใช้ลง stock PEAK + คิดกำไร + เป็นฐานมูลค่าสำแดงใบขน",None,F("434343",False,9),AL)
ws.row_dimensions[39].height=52
ph={"A39":("ลำดับ",MAROON),"B39":("ชื่อสินค้า (EN + TH)",MAROON),"F39":("จำนวน",MAROON),
 "G39":("RMB/PCS\nต้นทุนหยวน\nต่อชิ้น",ORANGE),"H39":("THB/PCS\nต้นทุนบาท\nต่อชิ้น",MAROON),
 "I39":("ต้นทุนรวม\n(บาท)",MAROON),"J39":("THB/PCS\nราคาขาย",BLUE),"K39":("ราคาขายรวม\n(บาท)",BLUE),
 "L39":("กำไร/ชิ้น\n(บาท)",GREEN),"M39":("กำไรรวม\n(บาท)",GREEN),"N39":("% กำไร",GREEN),
 "O39":("PO / แทร็คกิ้ง\nสั่งซื้อจีน",ORANGE),"S39":("หมายเหตุ\nต้นทุน",MAROON)}
merge(ws,"B39:E39"); merge(ws,"O39:R39"); merge(ws,"S39:X39")
for c,(t,fl) in ph.items(): put(ws,c,t,fl,F(WHITE,True,11))
for i in range(20):
    r=40+i; s=12+i  # pricing row r ↔ section-1 row s
    ws.row_dimensions[r].height=26
    merge(ws,f"B{r}:E{r}"); merge(ws,f"O{r}:R{r}"); merge(ws,f"S{r}:X{r}")
    put(ws,f"A{r}",f"=A{s}",LPINK2,F("434343"))
    put(ws,f"B{r}",f'=IF(G{s}="","",G{s}&"  "&J{s})',WHITE,F("434343"),AL)
    put(ws,f"F{r}",f"=M{s}",WHITE,F("434343"))
    put(ws,f"G{r}",None,YEL,F("000000",True)); ws[f"G{r}"].number_format="#,##0.0000"   # INPUT cost RMB
    put(ws,f"H{r}",f"=G{r}*$B$38",GREY,F("000000")); ws[f"H{r}"].number_format="#,##0.0000"
    put(ws,f"I{r}",f"=H{r}*F{r}",GREY,F("000000",True)); ws[f"I{r}"].number_format="#,##0.00"
    put(ws,f"J{r}",f"=P{s}",LBLUE,F("000000")); ws[f"J{r}"].number_format="#,##0.0000"
    put(ws,f"K{r}",f"=R{s}",LBLUE,F("000000",True)); ws[f"K{r}"].number_format="#,##0.00"
    put(ws,f"L{r}",f'=IF(G{r}="","",J{r}-H{r})',LGREEN,F("000000")); ws[f"L{r}"].number_format="#,##0.0000"
    put(ws,f"M{r}",f'=IF(G{r}="","",K{r}-I{r})',LGREEN,F("006100",True)); ws[f"M{r}"].number_format="#,##0.00"
    put(ws,f"N{r}",f'=IFERROR(M{r}/K{r},"")',LGREEN,F("006100")); ws[f"N{r}"].number_format="0.0%"
    put(ws,f"O{r}",None,WHITE,F("434343"),AL); put(ws,f"S{r}",None,WHITE,F("434343"),AL)
# pricing totals
ws.row_dimensions[60].height=30
merge(ws,"A60:F60"); put(ws,"A60","รวม",MAROON,F(WHITE,True,13),AL)
put(ws,"H60","ต้นทุน→",MAROON,F(WHITE,True,10),AR)
put(ws,"I60","=SUM(I40:I59)",MAROON,F(WHITE,True,12)); ws["I60"].number_format="#,##0.00"
put(ws,"J60","ขาย→",BLUE,F(WHITE,True,10),AR)
put(ws,"K60","=SUM(K40:K59)",BLUE,F(WHITE,True,12)); ws["K60"].number_format="#,##0.00"
put(ws,"L60","กำไร→",GREEN,F(WHITE,True,10),AR)
put(ws,"M60","=SUM(M40:M59)",GREEN,F(WHITE,True,12)); ws["M60"].number_format="#,##0.00"
put(ws,"N60",'=IFERROR(M60/K60,"")',GREEN,F(WHITE,True,12)); ws["N60"].number_format="0.0%"
print("section 2 done")

# ════════════════════════════════════════════════════════════════════
# SECTION 3 — DOCS (เจ้าหน้าที่ คีย์ใบขน) : NETBAY · มูลค่าสำแดง (default=cost, editable)
# ════════════════════════════════════════════════════════════════════
bar(ws,63,"A63:X63","3.  สำหรับเจ้าหน้าที่ คีย์ใบขน เท่านั้น  (NETBAY · ใบขนสินค้า)",BLUE)
ws.row_dimensions[64].height=40
merge(ws,"A64:B64"); put(ws,"A64","NETBAY",TEAL,F("000000",True,14))
put(ws,"C64","USD RATE",PINK,F("434343",True,11))
put(ws,"D64",31.5,CYAN,F("000000",True,14)); ws["D64"].number_format="0.00"
merge(ws,"E64:H64"); put(ws,"E64","ระบุใบอนุญาติ / Form (E·RCEP·D)",None,F("434343"),AL)
merge(ws,"I64:K64"); put(ws,"I64","ตู้ที่ลงใบขน",None,F("434343"),AL)
merge(ws,"L64:X64"); put(ws,"L64","⚠ มูลค่าสำแดง = ค่าเริ่มต้นดึงจาก 'ราคาต้นทุน' (Pricing) · DOCS แก้/ลดได้ตามแผนสำแดง (ช่องเหลือง) · USD RATE = เรทศุลกากรประจำเดือน (customs.go.th)",LORANGE,F("434343",False,9),AL)
ws.row_dimensions[65].height=52
dh={"A65":("ลำดับ",BLUE),"B65":("Traffic code\n(HS / พิกัด)",BLUE),"D65":("EN Description",BLUE),
 "H65":("TH Description",BLUE),"L65":("BRAND",BLUE),"M65":("Shipping Mark",BLUE),"O65":("Package\nCT",BLUE),
 "P65":("Qty\nInvoice",BLUE),"Q65":("Gross\nWeight",BLUE),"R65":("มูลค่าสำแดง\nบาท/หน่วย",ORANGE),
 "S65":("มูลค่าสำแดง\nรวม (บาท)",BLUE),"T65":("Price USD\n/ unit",BLUE),"U65":("Price USD\nรวม",BLUE),
 "V65":("Form / สำแดงชื่อ\n/ ใบอนุญาติ",BLUE)}
merge(ws,"B65:C65"); merge(ws,"D65:G65"); merge(ws,"H65:K65"); merge(ws,"M65:N65"); merge(ws,"V65:X65")
for c,(t,fl) in dh.items(): put(ws,c,t,fl,F(WHITE,True,10))
for i in range(20):
    r=66+i; s=12+i; pr=40+i
    ws.row_dimensions[r].height=24
    merge(ws,f"B{r}:C{r}"); merge(ws,f"D{r}:G{r}"); merge(ws,f"H{r}:K{r}"); merge(ws,f"M{r}:N{r}"); merge(ws,f"V{r}:X{r}")
    put(ws,f"A{r}",f"=A{s}",LBLUE,F("434343"))
    put(ws,f"B{r}",f"=C{s}",WHITE,F("434343"))
    put(ws,f"D{r}",f"=G{s}",WHITE,F("434343"),AL)
    put(ws,f"H{r}",f"=J{s}",WHITE,F("434343"),AL)
    put(ws,f"L{r}",f"=L{s}",WHITE,F("434343"))
    put(ws,f"M{r}","AXELRA (THAILAND) CO., LTD.",WHITE,F("434343",False,9),AL)
    put(ws,f"O{r}",f"=T{s}",WHITE,F("434343"))
    put(ws,f"P{r}",f"=M{s}",WHITE,F("434343"))
    put(ws,f"Q{r}",f"=U{s}",WHITE,F("434343")); ws[f"Q{r}"].number_format="#,##0.00"
    put(ws,f"R{r}",f'=IF(H{pr}="","",H{pr})',YEL,F("000000",True)); ws[f"R{r}"].number_format="#,##0.0000"  # สำแดง default=cost, editable
    put(ws,f"S{r}",f"=R{r}*P{r}",WHITE,F("000000")); ws[f"S{r}"].number_format="#,##0.00"
    put(ws,f"T{r}",f'=IFERROR(ROUND(R{r}/$D$64,2),"")',WHITE,F("000000")); ws[f"T{r}"].number_format="#,##0.00"
    put(ws,f"U{r}",f'=IFERROR(ROUND(S{r}/$D$64,2),"")',WHITE,F("000000")); ws[f"U{r}"].number_format="#,##0.00"
    put(ws,f"V{r}",None,WHITE,F("434343"),AL)
merge(ws,"A86:Q86"); put(ws,"A86","รวมมูลค่าสำแดง (ฐานคิดอากร + VAT ใบขน) →",BLUE,F(WHITE,True,11),AR)
put(ws,"R86",None,BLUE); put(ws,"S86","=SUM(S66:S85)",BLUE,F(WHITE,True,12)); ws["S86"].number_format="#,##0.00"
put(ws,"U86","=SUM(U66:U85)",BLUE,F(WHITE,True,12)); ws["U86"].number_format="#,##0.00"
print("section 3 done")

# ════════════════════════════════════════════════════════════════════
# SECTION 4 — ACCOUNT (เจ้าหน้าที่ บัญชี) : PEAK stock (cost) + ออกใบกำกับ (selling)
# ════════════════════════════════════════════════════════════════════
bar(ws,88,"A88:X88","4.  สำหรับเจ้าหน้าที่ บัญชี เท่านั้น  (PEAK · ลง stock ต้นทุน + ออกใบกำกับราคาขาย)",GREEN)
merge(ws,"A90:E90"); put(ws,"A90","ข้อมูลลูกค้า สำหรับออกใบกำกับ",BLUE,F(WHITE,True,12),AL)
acc=[("วันที่ออกใบกำกับภาษี","=R3"),("วันที่ ชำระเงิน","=R5"),("บุคคลฯ หรือ บริษัท","=C3"),
 ("เลขทะเบียน 13 หลัก","=C6"),("ชื่อสำหรับออกใบกำกับ","=C3"),("สำนักงานใหญ่ / สาขา","=C7"),
 ("ที่อยู่","=C4"),("ข้อมูลผู้ติดต่อ","=J3"),("เบอร์โทรศัพท์","=J4"),("Email","=J5"),
 ("ที่อยู่ส่งเอกสาร (ถ้ามี)",None),("โอนผ่านบัญชี","=C8")]
for i,(lab,val) in enumerate(acc):
    r=91+i; merge(ws,f"A{r}:C{r}"); merge(ws,f"D{r}:H{r}")
    put(ws,f"A{r}",lab,GREEN,F(WHITE,True,11),AL)
    put(ws,f"D{r}",val,LGREEN,F("000000"),AL)
# items (right): สินค้า/บริการ | จำนวน | ขาย/หน่วย | รวมขาย | ทุน/หน่วย | รวมทุน | ส่วนลด
merge(ws,"J90:X90"); put(ws,"J90","รายการ (ออกใบกำกับ = ราคาขาย+VAT7% · ลง stock PEAK = ราคาทุน)",GREEN,F(WHITE,True,12),AL)
ih={"I91":"ลำดับ","J91":"สินค้า / บริการ","N91":"จำนวน","O91":"ราคาขาย\n/หน่วย","P91":"รวมขาย",
 "R91":"ราคาทุน\n/หน่วย (PEAK)","S91":"รวมทุน","U91":"ส่วนลด\n/หน่วย"}
ws.row_dimensions[91].height=42
merge(ws,"J91:M91"); merge(ws,"P91:Q91"); merge(ws,"S91:T91"); merge(ws,"U91:X91")
for c,t in ih.items(): put(ws,c,t,BLUE,F(WHITE,True,10))
for i in range(20):
    r=92+i; s=12+i; pr=40+i
    ws.row_dimensions[r].height=22
    merge(ws,f"J{r}:M{r}"); merge(ws,f"P{r}:Q{r}"); merge(ws,f"S{r}:T{r}"); merge(ws,f"U{r}:X{r}")
    put(ws,f"I{r}",f"=A{s}",LGREEN,F("434343"))
    put(ws,f"J{r}",f'=IF(G{s}="","",G{s}&"  "&J{s})',WHITE,F("434343"),AL)
    put(ws,f"N{r}",f"=M{s}",WHITE,F("434343"))
    put(ws,f"O{r}",f"=P{s}",WHITE,F("000000")); ws[f"O{r}"].number_format="#,##0.0000"
    put(ws,f"P{r}",f"=R{s}",WHITE,F("000000")); ws[f"P{r}"].number_format="#,##0.00"
    put(ws,f"R{r}",f"=H{pr}",LGREEN,F("000000")); ws[f"R{r}"].number_format="#,##0.0000"
    put(ws,f"S{r}",f"=I{pr}",LGREEN,F("000000")); ws[f"S{r}"].number_format="#,##0.00"
    put(ws,f"U{r}",None,WHITE,F("434343"))
r=112
merge(ws,f"J{r}:O{r}"); put(ws,f"J{r}","รวมขาย (ฐาน VAT) →",GREEN,F(WHITE,True,11),AR)
put(ws,f"P{r}","=SUM(P92:P111)",GREEN,F(WHITE,True,12)); ws[f"P{r}"].number_format="#,##0.00"
merge(ws,f"Q{r}:R{r}"); put(ws,f"Q{r}","ทุน →",GREEN,F(WHITE,True,11),AR)
put(ws,f"S{r}","=SUM(S92:S111)",GREEN,F(WHITE,True,12)); ws[f"S{r}"].number_format="#,##0.00"
r=113; merge(ws,f"J{r}:O{r}"); put(ws,f"J{r}","VAT 7% →",GREEN,F(WHITE,True,11),AR)
put(ws,f"P{r}","=ROUND(P112*0.07,2)",GREEN,F(WHITE,True,12)); ws[f"P{r}"].number_format="#,##0.00"
merge(ws,f"Q{r}:R{r}"); put(ws,f"Q{r}","กำไร →",GREEN,F(WHITE,True,11),AR)
put(ws,f"S{r}","=P112-S112",GREEN,F(WHITE,True,12)); ws[f"S{r}"].number_format="#,##0.00"
r=114; merge(ws,f"J{r}:O{r}"); put(ws,f"J{r}","รวมทั้งสิ้น (ลูกค้าชำระ) →",GREEN,F(WHITE,True,12),AR)
merge(ws,f"P{r}:Q{r}"); put(ws,f"P{r}","=P112+P113",GREEN,F(WHITE,True,13)); ws[f"P{r}"].number_format="#,##0.00"
print("section 4 done")

# print area / freeze
ws.print_area = "A1:X114"
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1; ws.sheet_properties.pageSetUpPr = None

# ════════════════════════════════════════════════════════════════════
# INSTRUCTIONS sheet
# ════════════════════════════════════════════════════════════════════
ins = wb.create_sheet("วิธีใช้ (อ่านก่อน)")
ins.sheet_view.showGridLines=False
ins.column_dimensions["A"].width=3; ins.column_dimensions["B"].width=120
rows = [
 ("H","ฟอร์มใบกำกับภาษี Pacred (CARGO) — 4 ขั้นตอน : CS → Pricing → Docs → Account"),
 ("T",""),
 ("S","ภาพรวม : งาน CARGO = งาน Freight LCL รูปแบบหนึ่ง ที่ Pacred ออก 'ใบขนรวม' ในนามชิปปิ้ง (Pacred/AXELRA) ลูกค้าจึงเห็นแค่ใบกำกับภาษี ไม่เห็นใบขน"),
 ("S","ถ้าลูกค้าอยากได้ใบขนในชื่อตัวเอง → กลายเป็นงาน FREIGHT LCL (อีกฟอร์ม). ฟอร์มนี้ = เคสลูกค้า 'เอาเอกสาร' (จ่าย VAT7% รับใบกำกับ)."),
 ("T",""),
 ("K","⭐ ตัวเลข 3 ก้อนที่ห้ามสับสน (หัวใจของฟอร์ม)"),
 ("B","• ราคาขาย (SELLING) — CS กรอก (ช่อง 1) · = ราคาที่ลูกค้าจ่าย · เป็นฐานคิด VAT 7% · ออกใบกำกับด้วยราคานี้"),
 ("B","• ราคาต้นทุน (COST) — Pricing กรอก (ช่อง 2 · ช่องเหลือง RMB/PCS) · = ราคาที่ Pacred ซื้อจริงจากจีน · ใช้ลง stock PEAK + คิดกำไร"),
 ("B","• มูลค่าสำแดง (DECLARED) — Docs กรอก (ช่อง 3 · ช่องเหลือง) · = ราคาที่สำแดงในใบขน · ค่าเริ่มต้นดึงจากต้นทุน แต่ Docs ปรับลด/แก้ได้ตามแผนสำแดง"),
 ("T",""),
 ("1","ขั้นที่ 1 — CS : ทำสำเนาฟอร์ม → กรอกข้อมูลลูกค้า + รายการสินค้า (รูป/HS/ชื่อ/จำนวน) + RMB/PCS 'ราคาขาย' (THB/PCS, ยอดสุทธิ, VAT7% คำนวณอัตโนมัติ) → แคปเฟิร์มลูกค้า → แนบสลิปไทย (มุมขวาบน)"),
 ("2","ขั้นที่ 2 — PRICING (ส่วนที่เพิ่มใหม่) : กรอก RMB/PCS 'ต้นทุน' (ช่องเหลือง) + เรทหยวนต้นทุน → ระบบคิด ต้นทุนบาท/รวม/กำไร/%กำไร อัตโนมัติ → แนบสลิปจีน (มุมขวาบน). ต้นทุนนี้ส่งต่อให้ Docs (ฐานสำแดง) + Account (ลง stock PEAK)"),
 ("3","ขั้นที่ 3 — DOCS : ใส่ USD RATE (เรทศุลกากรประจำเดือน) + Form E/RCEP/ใบอนุญาติ → ตรวจ/แก้ 'มูลค่าสำแดง' (ดึงจากต้นทุน แก้ได้) → ทำ Invoice/Packing List + ขอ Form E → คีย์ใบขนใน NETBAY → ยิงใบขน → ตั้งเบิกตัดภาษีกับบัญชี"),
 ("4","ขั้นที่ 4 — ACCOUNT : เอาใบขนที่ปิดชุดแล้ว → ลงรายการ 'ราคาทุน' เข้า stock PEAK → ออกใบกำกับภาษี 'ราคาขาย' (+VAT7%) ใน PEAK ส่งลูกค้า → ปิดจบชุดงาน"),
 ("T",""),
 ("N","หมายเหตุที่ต้องเติมเอง (ไม่มีในแหล่งข้อมูล) : รหัสบัญชี PEAK + เลข/ฟิลด์ NETBAY ให้ฝ่ายบัญชี (NAT) + Docs กรอกตามระบบจริง"),
 ("N","อ้างอิงมาร์กอัป (เฟรท/ขนส่ง 30/25/20/15/10% ตามขนาดลูกค้า) · 1 CBM=300KG (cargo) · Form E/RCEP → อากร 0% (ACFTA) · ทศนิยมไม่เกิน 4 ตำแหน่ง"),
]
ir=2
styles={"H":(BLUE,WHITE,True,16),"K":(MAROON,WHITE,True,13),"S":(LORANGE,"434343",False,11),
 "1":(LPINK2,"434343",False,11),"2":(LORANGE,"990000",True,11),"3":(LBLUE,"434343",False,11),
 "4":(LGREEN,"434343",False,11),"N":(GREY,"434343",False,10),"B":(WHITE,"434343",False,11),"T":(WHITE,"434343",False,8)}
for tag,txt in rows:
    fl,fc,b,sz = styles[tag]
    ins.row_dimensions[ir].height = 30 if tag in("H","K") else (10 if tag=="T" else 22)
    put(ins,f"B{ir}",txt,fl if tag!="B" and tag!="T" else None,F(fc,b,sz),AL,border=False)
    if tag=="H": merge(ins,f"A{ir}:B{ir}")
    ir+=1
wb.move_sheet("วิธีใช้ (อ่านก่อน)", -(len(wb.sheetnames)-1))  # move to front

wb.save("/tmp/PACRED-ใบกำกับภาษี-form.xlsx")
print("SAVED full form")
